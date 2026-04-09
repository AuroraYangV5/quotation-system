#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报价生成系统 Web后端
FastAPI应用入口
"""

import os
import json
import urllib
import base64
import tempfile
import uuid
import xlrd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional

app = FastAPI(title="报价自动生成系统")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# 创建临时目录存储上传文件和生成文件
TEMP_DIR = tempfile.gettempdir()
os.makedirs(TEMP_DIR, exist_ok=True)

# GLM API地址
GLM_API_URL = "https://open.bigmodel.cn/api/paas/v4/chat/completions"


# Pydantic模型
class GeneratedItem(BaseModel):
    sheetName: str
    spec: str
    ruleName: str
    originalPrice: float
    profitPercent: float

    model_config = {"extra": "allow"}  # 允许额外的动态字段


class SheetInfo(BaseModel):
    sheetName: str
    tableTitle: Optional[str]
    date: Optional[str]


class GenerateRequest(BaseModel):
    items: List[dict]  # 支持动态自定义字段
    sheetInfos: List[SheetInfo]

    model_config = {"extra": "allow"}


# 解析结果模型
class ParseItem(BaseModel):
    id: str
    sheetName: str
    tableIndex: int
    spec: str
    ruleName: str
    originalPrice: float


class ParseTable(BaseModel):
    sheetName: str
    tableTitle: Optional[str]
    date: Optional[str]
    rules: List[str]
    items: List[ParseItem]


class ParseResponse(BaseModel):
    success: bool
    message: str
    data: Optional[dict]


@app.post("/api/parse", response_model=ParseResponse)
async def parse_excel(file: UploadFile = File(...)):
    """
    解析Excel文件，返回结构化数据供前端预览编辑
    """
    filename = file.filename
    if not filename:
        return ParseResponse(success=False, message="未上传文件", data=None)

    ext = os.path.splitext(filename)[1].lower()
    if ext not in ['.xls', '.xlsx']:
        return ParseResponse(success=False, message="只支持 .xls .xlsx 格式的Excel文件", data=None)

    try:
        # 生成唯一ID避免文件冲突
        file_id = str(uuid.uuid4())[:8]

        # 保存上传的文件
        input_path = os.path.join(TEMP_DIR, f"input_{file_id}{ext}")
        with open(input_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # 使用xlrd解析
        wb = xlrd.open_workbook(input_path)

        tables: List[ParseTable] = []
        total_items = 0

        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            title = None
            date = None

            # 找标题和日期
            for row_idx in range(min(3, sheet.nrows)):
                for col_idx in range(min(5, sheet.ncols)):
                    cell_val = sheet.cell_value(row_idx, col_idx)
                    if cell_val and '价格' in str(cell_val) and not title:
                        title = str(cell_val).strip()
                    if cell_val and str(cell_val).count('.') >= 1 and not date:
                        try:
                            float(str(cell_val))
                            date = str(cell_val)
                        except:
                            pass

            if not title:
                title = sheet_name

            # 检测表格区域
            spec_found = False
            tables_in_sheet = []
            for row_idx in range(0, min(10, sheet.nrows)):
                cell_val = str(sheet.cell_value(row_idx, 0)).strip()
                if '规格' in cell_val or cell_val == '规格':
                    # 找到表头
                    rules = []
                    for col_idx in range(1, sheet.ncols):
                        cell_val = str(sheet.cell_value(row_idx, col_idx)).strip()
                        if cell_val and cell_val != '':
                            if '地址' in cell_val or '电话' in cell_val:
                                continue
                            rules.append(cell_val)

                    if len(rules) > 0:
                        tables_in_sheet.append({
                            'start_row': row_idx,
                            'spec_col': 0,
                            'rules': rules,
                            'rule_col_start': 1
                        })
                        spec_found = True
                    continue

                # 处理规格不在第一列的情况
                if not spec_found:
                    for col_idx in range(1, min(10, sheet.ncols)):
                        cell_val = str(sheet.cell_value(row_idx, col_idx)).strip()
                        if '规格' in cell_val or cell_val == '规格':
                            rules = []
                            for c_idx in range(0, sheet.ncols):
                                if c_idx == col_idx:
                                    continue
                                cell_val = str(sheet.cell_value(row_idx, c_idx)).strip()
                                if cell_val and cell_val != '':
                                    if '地址' in cell_val or '电话' in cell_val:
                                        continue
                                    rules.append(cell_val)
                            if len(rules) > 0:
                                tables_in_sheet.append({
                                    'start_row': row_idx,
                                    'spec_col': col_idx,
                                    'rules': rules,
                                    'rule_col_start': 0
                                })
                                spec_found = True
                            break

            # 如果没找到规格表头，默认检测
            if not spec_found and len(tables_in_sheet) == 0:
                for row_idx in range(0, min(5, sheet.nrows)):
                    has_content = False
                    for col_idx in range(1, sheet.ncols):
                        if sheet.cell_value(row_idx, col_idx):
                            has_content = True
                            break
                    if has_content:
                        rules = []
                        for col_idx in range(1, sheet.ncols):
                            cell_val = str(sheet.cell_value(row_idx, col_idx)).strip()
                            if cell_val and cell_val != '':
                                if '地址' in cell_val or '电话' in cell_val:
                                    continue
                                rules.append(cell_val)
                        if len(rules) > 0:
                            tables_in_sheet.append({
                                'start_row': row_idx,
                                'spec_col': 0,
                                'rules': rules,
                                'rule_col_start': 1
                            })
                        break

            # 解析每个表格数据，同一个sheet中的多个table合并到一个sheet
            all_items_in_sheet: List[ParseItem] = []
            for table_idx, table_info in enumerate(tables_in_sheet):
                start_data_row = table_info['start_row'] + 1
                spec_col = table_info['spec_col']
                rules = table_info['rules']
                rule_col_start = table_info['rule_col_start']

                for row_idx in range(start_data_row, sheet.nrows):
                    # 跳过空行
                    row_is_empty = True
                    for col_idx in range(sheet.ncols):
                        if sheet.cell_value(row_idx, col_idx):
                            row_is_empty = False
                            break
                    if row_is_empty:
                        continue

                    # 获取规格
                    spec_cell = sheet.cell_value(row_idx, spec_col)
                    spec_str = str(spec_cell).strip()
                    if '地址' in spec_str or '电话' in spec_str or len(spec_str) > 30:
                        continue
                    if spec_str == '':
                        continue

                    # 获取每个规则的价格
                    for i, rule in enumerate(rules):
                        col_idx = rule_col_start + i
                        if col_idx >= sheet.ncols:
                            continue
                        cell_val = sheet.cell_value(row_idx, col_idx)
                        try:
                            if isinstance(cell_val, float):
                                price = cell_val
                            else:
                                price = float(str(cell_val)) if str(cell_val).strip() else 0
                        except:
                            price = 0
                        if price <= 0:
                            continue

                        item_id = f"{sheet_name}-{table_idx}-{row_idx}-{i}"
                        all_items_in_sheet.append(ParseItem(
                            id=item_id,
                            sheetName=sheet_name,
                            tableIndex=table_idx,
                            spec=spec_str,
                            ruleName=rule,
                            originalPrice=price
                        ))
                        total_items += 1

            # 如果当前sheet有数据，添加为一个sheet（合并多个表格）
            if len(all_items_in_sheet) > 0:
                # 收集所有不重复的rules（其实不需要，因为items已经每个都带ruleName了）
                # 前端只需要items，按sheet分组就行
                tables.append(ParseTable(
                    sheetName=sheet_name,
                    tableTitle=title,
                    date=date,
                    rules=[],
                    items=all_items_in_sheet
                ))

        if len(tables) == 0 or total_items == 0:
            return ParseResponse(
                success=False,
                message="未能解析到任何价格数据，请检查文件格式",
                data=None
            )

        sheets = list(set(t.sheetName for t in tables))

        return ParseResponse(
            success=True,
            message=f"解析成功，共 {len(tables)} 个表格，{total_items} 个价格项",
            data={
                'sheets': sheets,
                'tables': [t.dict() for t in tables],
                'totalItems': total_items
            }
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return ParseResponse(success=False, message=f"解析失败: {str(e)}", data=None)


@app.post("/api/ocr-parse")
async def ocr_parse(file: UploadFile = File(...)):
    """
    图片OCR识别，从图片中提取规格和价格
    使用GLM-4V-Flash视觉API识别，返回格式与 /api/parse 完全相同
    """
    try:
        # 保存临时文件
        file_id = str(uuid.uuid4())[:8]
        temp_path = os.path.join(TEMP_DIR, f"ocr_{file_id}_{file.filename}")

        with open(temp_path, "wb") as buffer:
            buffer.write(await file.read())

        # 调用GLM-4V-Flash解析
        from parse_image_glm import parse_image
        result = parse_image(temp_path)

        # 删除临时文件
        os.remove(temp_path)

        return result
    except Exception as e:
        import traceback
        traceback.print_exc()
        return ParseResponse(success=False, message=f"OCR识别失败: {str(e)}", data=None)


@app.post("/api/parse-by-glm-preview")
async def convert_excel_to_preview(file: UploadFile = File(...)):
    """
    将Excel转换为图片预览（供用户检查转换效果）
    返回转换后的图片URL列表
    """
    filename = file.filename
    if not filename:
        return {"success": False, "message": "未上传文件", "data": None}

    ext = os.path.splitext(filename)[1].lower()
    if ext not in ['.xls', '.xlsx', '.pdf']:
        return {"success": False, "message": "只支持 .xls .xlsx .pdf 格式", "data": None}

    try:
        # 生成唯一ID避免文件冲突
        file_id = str(uuid.uuid4())[:8]

        # 保存上传的Excel文件
        input_path = os.path.join(TEMP_DIR, f"glm_input_{file_id}{ext}")
        with open(input_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # 1. 将Excel/PDF转换为图片
        from excel_to_image import ExcelToImageConverter, get_excel_reader
        image_paths = []
        page_names = []

        if ext == '.pdf':
            # PDF转换：每一页一张图片
            from pdf2image import convert_from_path
            pages = convert_from_path(input_path, dpi=150)
            for idx, page in enumerate(pages):
                page_name = f"第{idx+1}页"
                img_filename = f"glm_preview_{file_id}_{idx}_page{idx+1}.png"
                img_path = os.path.join(TEMP_DIR, img_filename)
                page.save(img_path, 'PNG')
                image_paths.append(f"/api/glm-preview/{img_filename}")
                page_names.append(page_name)
        else:
            # Excel转换
            converter = ExcelToImageConverter()
            sheet_names_raw, _ = get_excel_reader(input_path)
            for idx, sheet_name in enumerate(sheet_names_raw):
                converter = ExcelToImageConverter()
                img = converter.convert_sheet(input_path, idx)
                img_filename = f"glm_preview_{file_id}_{idx}_{sheet_name}.png"
                img_filename = img_filename.replace('/', '_').replace('\\', '_')
                img_path = os.path.join(TEMP_DIR, img_filename)
                img.save(img_path, 'PNG')
                image_paths.append(f"/api/glm-preview/{img_filename}")
                page_names.append(sheet_name)

        os.remove(input_path)

        return {
            "success": True,
            "message": f"转换成功，共 {len(image_paths)} 页/Sheet",
            "data": {
                "file_id": file_id,
                "images": [
                    {"url": url, "sheet_name": name}
                    for url, name in zip(image_paths, page_names)
                ]
            }
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "message": f"转换失败: {str(e)}", "data": None}


@app.get("/api/glm-preview/{filename}")
async def get_glm_preview(filename: str):
    """获取转换后的预览图片"""
    file_path = os.path.join(TEMP_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="图片不存在")
    return FileResponse(file_path, media_type="image/png")


from fastapi import Query

@app.post("/api/parse-by-glm")
async def parse_excel_by_glm(
    file: UploadFile = File(...),
    selected_sheets: str = Query(None, description="选中的Sheet索引，逗号分隔"),
    fields: str = Query(None, description="自定义提取字段，JSON格式")
):
    """
    使用GLM-4V-Flash解析Excel文件
    流程：Excel -> 转换为图片（每个Sheet一张）-> 一次性把所有图片传给GLM -> 返回完整JSON
    保留原有 /api/parse 作为备用方案
    支持selected_sheets参数，只解析选中的Sheet索引
    """
    filename = file.filename
    if not filename:
        return ParseResponse(success=False, message="未上传文件", data=None)

    ext = os.path.splitext(filename)[1].lower()
    if ext not in ['.xls', '.xlsx', '.pdf']:
        return ParseResponse(success=False, message="只支持 .xls .xlsx .pdf 格式", data=None)

    # 解析选中的Sheet/Page索引
    selected_indices = None
    if selected_sheets:
        try:
            selected_indices = set(int(i) for i in selected_sheets.split(','))
        except:
            pass

    # 解析自定义提取字段
    import json
    configurable_fields = None
    if fields:
        try:
            configurable_fields = json.loads(urllib.parse.unquote(fields))
        except:
            pass

    try:
        # 生成唯一ID避免文件冲突
        file_id = str(uuid.uuid4())[:8]
        api_key = os.getenv("ZHIPU_API_KEY")
        if not api_key:
            # 如果环境变量未设置，尝试从硬编码读取（仅供开发）
            api_key = '5b34e11121f5438d9284f04d8a69ae08.3ZFZSFLtScezhqUB'
        if not api_key:
            return {
                'success': False,
                'message': '未配置 ZHIPU_API_KEY 环境变量，请先设置智谱AI API Key',
                'data': None
            }

        # 保存上传的文件
        input_path = os.path.join(TEMP_DIR, f"glm_input_{file_id}{ext}")
        with open(input_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # 1. 将选中的页面转换为图片，并编码为base64
        encoded_images = []
        temp_image_paths = []

        if ext == '.pdf':
            # PDF转换：每一页一张图片
            from pdf2image import convert_from_path
            pages = convert_from_path(input_path, dpi=150)
            for idx, page in enumerate(pages):
                # 如果指定了选中的页码，只处理选中的
                if selected_indices is not None and idx not in selected_indices:
                    continue
                page_name = f"第{idx+1}页"
                # 保存到临时文件
                img_filename = f"glm_tmp_{file_id}_{idx}.png"
                img_path = os.path.join(TEMP_DIR, img_filename)
                page.save(img_path, 'PNG')
                temp_image_paths.append(img_path)
                # 编码为base64
                with open(img_path, "rb") as f:
                    base64_img = base64.b64encode(f.read()).decode('utf-8')
                    encoded_images.append({
                        'sheet_name': page_name,
                        'base64': base64_img
                    })
        else:
            # Excel转换
            from excel_to_image import ExcelToImageConverter, get_excel_reader
            converter = ExcelToImageConverter()
            sheet_names, _ = get_excel_reader(input_path)
            for idx, sheet_name in enumerate(sheet_names):
                # 如果指定了选中的Sheet，只处理选中的
                if selected_indices is not None and idx not in selected_indices:
                    continue
                converter = ExcelToImageConverter()
                img = converter.convert_sheet(input_path, idx)
                # 保存到临时文件
                img_filename = f"glm_tmp_{file_id}_{idx}.png"
                img_path = os.path.join(TEMP_DIR, img_filename)
                img.save(img_path, 'PNG')
                temp_image_paths.append(img_path)
                # 编码为base64
                with open(img_path, "rb") as f:
                    base64_img = base64.b64encode(f.read()).decode('utf-8')
                    encoded_images.append({
                        'sheet_name': sheet_name,
                        'base64': base64_img
                    })

        # 2. 逐张图片串行调用GLM API，带指数退避重试应对频率限制
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }

        # 为单张图片构建prompt
        if configurable_fields and len(configurable_fields) > 0:
            # 使用用户自定义字段
            field_desc = "\n".join([
                f"  - {f['key']}: {f['name']} - {f['description']}"
                for f in configurable_fields
            ])
            example_items = ",\n    ".join([
                f'      "{f["key"]}": "{f["name"]}字符串"'
                for f in configurable_fields
            ])
            single_prompt = f"""
你需要识别这张价格表图片中的所有商品数据，按照以下指定字段提取：
{field_desc}

输出要求：
- 所有字段**必须都是字符串类型**
- 必须只返回一个完整的JSON，格式如下：
{{
  "items": [
    {{
{example_items}
    }}
  ]
}}
"""
        else:
            # 默认字段
            single_prompt = """
请识别这张价格表图片中的所有商品规格和原价。

输出要求：
1. 提取表格中的每一行数据
2. 第一列通常是规格(如 76, 89, 114 等)，其他列是不同类型(如 卡箍, 弯头, 三通 等)和价格
3. 每行每个规格-类型-价格组合输出一条记录
4. **所有字段都必须是字符串类型**，包括价格
5. **必须只返回一个完整的JSON**，格式如下：
{
  "items": [
    {
      "spec": "规格名称",
      "ruleName": "类型名称",
      "originalPrice": "价格字符串"
    }
  ]
}
"""

        prompt = single_prompt + """

重要提醒：
- 如果单元格为空，跳过不输出
- 价格只保留数字，不要带货币符号、空格等，但必须输出为字符串，不能是数字类型
- **只输出JSON，绝对不要输出任何其他解释文字、markdown标记、反引号**
- **确保JSON格式正确，不要有多余逗号，不要有多个JSON对象**
- 所有数据放在一个items数组里

最后警告：
- 你的输出 **必须** 仅仅是纯粹的JSON
- 不允许有任何额外的文字、说明、问候、标记
- 不允许用```json 和 ```包裹
- 直接输出JSON，说完就结束
"""

        # 逐张图片处理，串行调用
        import re
        import asyncio

        all_output_items = []
        all_errors = []
        tables = []
        max_retries = 3  # 最大重试次数
        processed_sheets = []

        for img_info in encoded_images:
            sheet_name = img_info['sheet_name']
            base64_img = img_info['base64']
            processed_sheets.append(sheet_name)

            print(f"正在处理: {sheet_name}")

            # 指数退避重试
            success = False
            for retry in range(max_retries):
                # 构建单张图片的请求
                content = [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64_img}"
                        }
                    }
                ]

                payload = {
                    "model": "glm-4.1v-thinking-flash",
                    "max_tokens": 16384,
                    "messages": [
                        {
                            "role": "user",
                            "content": content
                        }
                    ]
                }

                response = requests.post(GLM_API_URL, headers=headers, json=payload, timeout=300)

                # 处理限流 - 退避重试
                if response.status_code == 429:
                    wait_time = 2 ** retry  # 指数退避: 1s, 2s, 4s
                    print(f"{sheet_name} 触发限流，等待 {wait_time} 秒后重试...")
                    await asyncio.sleep(wait_time)
                    continue

                # 其他错误也可能重试
                if response.status_code >= 500:
                    wait_time = 2 ** retry
                    print(f"{sheet_name} 服务错误 {response.status_code}，等待 {wait_time} 秒后重试...")
                    await asyncio.sleep(wait_time)
                    continue

                # 成功收到响应
                break

            # 处理响应结果
            if response.status_code == 429:
                # 多次重试仍然限流
                all_errors.append(f"{sheet_name}: 请求频率过高，重试后仍然失败")
                continue

            if response.status_code != 200:
                all_errors.append(f"{sheet_name}: HTTP {response.status_code}")
                continue

            try:
                result = response.json()
            except Exception as e:
                all_errors.append(f"{sheet_name}: JSON解析失败: {str(e)}")
                continue

            # 提取回复内容
            try:
                content = result['choices'][0]['message']['content']
            except Exception as e:
                all_errors.append(f"{sheet_name}: 响应格式错误: {str(e)}")
                continue

            # 打印原始内容到控制台，方便调试
            print("\n" + "="*60)
            print(f"GLM返回 [{sheet_name}]:")
            print(content)
            print("="*60 + "\n")

            # 提取并解析JSON
            # 第一步：提取出JSON部分 - 找从第一个 { 到最后一个 }
            start = content.find('{')
            end = content.rfind('}') + 1
            if start >= 0 and end > start:
                json_candidate = content[start:end]
            else:
                all_errors.append(f"{sheet_name}: GLM未返回JSON格式内容")
                continue

            # 第二步：修复常见的JSON格式错误
            fixed = json_candidate
            # 1. 移除行尾注释（// ...）
            fixed = re.sub(r'//.*$', '', fixed, flags=re.MULTILINE)
            # 2. 移除多行注释（/* ... */）
            fixed = re.sub(r'/\*.*?\*/', '', fixed, flags=re.DOTALL)
            # 3. 修复 trailing commas: ", }" → " }" 和 ", ]" → " ]"
            fixed = re.sub(r',\s*([\]}])', r'\1', fixed)
            # 4. 修复多行情况下的 trailing commas
            fixed = re.sub(r',\s*\n\s*([\]}])', r'\n\1', fixed)
            # 5. 修复单引号改双引号
            fixed = re.sub(r"'([^']*)':", r'"\1":', fixed)
            # 6. 移除空行
            fixed = re.sub(r'\n\s*\n', r'\n', fixed)
            # 7. 如果有多个JSON对象，尝试合并items
            if '}{' in fixed:
                if 'items' in fixed:
                    fixed = re.sub(r'\}\s*\{', ',', fixed)

            # 压缩空格
            compact = re.sub(r'\s+', ' ', fixed)

            # 尝试解析
            try:
                data = json.loads(fixed)
            except json.JSONDecodeError as e1:
                try:
                    data = json.loads(compact)
                except json.JSONDecodeError as e2:
                    all_errors.append(f"{sheet_name}: 格式不对，请减少文件大小后重试")
                    continue

            items = data.get('items', [])
            if not items:
                all_errors.append(f"{sheet_name}: 未识别出任何商品数据")
                continue

            # 转换为统一格式，保留所有动态字段
            sheet_output_items = []
            for item in items:
                try:
                    # 提取必填的核心字段
                    spec = ''
                    original_price = 0
                    rule_name = ''

                    # 尝试从item中提取核心字段，如果不存在留空
                    for key in item.keys():
                        if key in ['spec', '规格', 'name']:
                            spec = str(item.get(key, '')).strip()
                        if key in ['originalPrice', '原价', 'price', '价格']:
                            price_val = item.get(key, '0')
                            if isinstance(price_val, str):
                                price_str = ''.join(c for c in price_val if c.isdigit() or c == '.')
                                original_price = float(price_str) if price_str else 0
                            else:
                                original_price = float(price_val) if price_val else 0
                        if key in ['ruleName', '类型', 'rule', '分类']:
                            rule_name = str(item.get(key, '')).strip()

                    # 保留所有原始字段，添加id等固定字段
                    output_item = dict(item)
                    output_item['id'] = str(uuid.uuid4())[:8]
                    output_item['sheetName'] = sheet_name
                    output_item['tableIndex'] = 0
                    # 确保核心字段存在
                    if 'spec' not in output_item and spec:
                        output_item['spec'] = spec
                    if 'originalPrice' not in output_item:
                        output_item['originalPrice'] = original_price
                    if 'ruleName' not in output_item and rule_name:
                        output_item['ruleName'] = rule_name

                    # 检查必须的spec和price
                    spec_val = output_item.get('spec', '')
                    price_val = output_item.get('originalPrice', 0)
                    if not spec_val:
                        continue
                    if isinstance(price_val, str):
                        price_str = ''.join(c for c in str(price_val) if c.isdigit() or c == '.')
                        output_item['originalPrice'] = float(price_str) if price_str else 0
                    if output_item['originalPrice'] <= 0:
                        continue

                    sheet_output_items.append(output_item)
                    all_output_items.append(output_item)
                except Exception:
                    continue

            # 当前Sheet处理成功，直接为这个Sheet创建一个table
            if len(sheet_output_items) > 0:
                sheet_rules = list(set(item['ruleName'] for item in sheet_output_items if item['ruleName']))
                if not sheet_rules:
                    sheet_rules = ["价格"]
                tables.append({
                    'sheetName': sheet_name,
                    'tableTitle': f'GLM AI识别 {sheet_name}',
                    'date': None,
                    'rules': sheet_rules,
                    'items': sheet_output_items
                })

            # 处理完一张图片，稍微暂停一下避免触发限流
            if len(encoded_images) > 1:
                await asyncio.sleep(0.5)

        # 所有图片处理完成，检查是否有任何结果
        if len(tables) == 0 or len(all_output_items) == 0:
            # 全部失败
            # 清理临时文件
            for img_path in temp_image_paths:
                os.remove(img_path)
            os.remove(input_path)
            error_msg = "; ".join(all_errors)
            return ParseResponse(
                success=False,
                message=f'全部处理失败。错误: {error_msg}',
                data=None
            )

        # 清理临时文件
        for img_path in temp_image_paths:
            os.remove(img_path)
        os.remove(input_path)

        # 构建返回信息
        total = len(all_output_items)
        sheet_names = [table['sheetName'] for table in tables]
        msg = f'GLM识别成功，共 {len(tables)} 个表格，{total} 个价格项'
        if all_errors:
            msg += f' ({len(all_errors)} 页处理失败)'

        # 返回结果
        return ParseResponse(
            success=True,
            message=msg,
            data={
                'sheets': sheet_names,
                'tables': tables,
                'totalItems': total
            }
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return ParseResponse(success=False, message=f"GLM解析失败: {str(e)}", data=None)


@app.post("/api/generate")
async def generate_excel(request: GenerateRequest):
    """
    根据前端编辑后的价格数据生成Excel文件并返回下载
    支持动态自定义字段，每个item占一行，包含所有字段
    """
    try:
        if len(request.items) == 0:
            raise HTTPException(status_code=400, detail="没有商品数据")

        # 生成输出文件
        file_id = str(uuid.uuid4())[:8]
        output_path = os.path.join(TEMP_DIR, f"报价表_{file_id}.xlsx")
        output_filename = "报价表.xlsx"

        # 创建新工作簿
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        # 样式定义
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        title_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        alt_row_fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')

        # 按sheet分组
        from collections import defaultdict
        items_by_sheet: defaultdict[str, List[dict]] = defaultdict(list)
        # 将items转为dict并按sheet分组
        for item in request.items:
            # Pydantic v2 方式获取所有字段（包括额外字段）
            if hasattr(item, 'model_dump'):
                item_dict = item.model_dump()
            elif isinstance(item, dict):
                item_dict = item
            else:
                item_dict = item.__dict__
                if '_state' in item_dict:
                    del item_dict['_state']
            sheet_name = item_dict.get('sheetName', '报价')
            items_by_sheet[sheet_name].append(item_dict)

        sheet_info_map = {si.sheetName: si for si in request.sheetInfos}

        # 固定字段（排除不导出的字段）
        # 排除：id、sheetName、tableIndex、选择设置利润、选择导出、原价、利润率
        exclude_fields = ['id', 'sheetName', 'tableIndex', 'selectedForBatch', 'selectedForExport', 'originalPrice', 'profitPercent']

        for sheet_name, items in items_by_sheet.items():
            ws = wb.create_sheet(title=sheet_name)
            sheet_info = sheet_info_map.get(sheet_name)
            title = sheet_info.tableTitle if sheet_info and sheet_info.tableTitle else sheet_name
            date = sheet_info.date if sheet_info else None

            current_row = 1

            # 收集所有列名：从所有items中收集所有不重复的字段
            all_columns = set()
            for item in items:
                for key in item.keys():
                    if key not in exclude_fields:
                        all_columns.add(key)
            # 固定列顺序：spec, ruleName, originalPrice, profitPercent, calculatedPrice，然后是其他自定义字段
            fixed_order = ['spec', 'ruleName', 'originalPrice', 'profitPercent', 'calculatedPrice']
            columns = []
            # 先加固定顺序的列（如果存在）
            for col in fixed_order:
                if col in all_columns:
                    columns.append(col)
                    all_columns.discard(col)
            # 再加剩余的自定义字段，按字母排序
            for col in sorted(all_columns):
                if col not in exclude_fields:
                    columns.append(col)

            total_cols = len(columns)
            end_col_letter = get_column_letter(total_cols)

            # 标题
            ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = f"{title}"
            cell.font = title_font
            cell.alignment = center_align
            cell.fill = title_fill
            ws.row_dimensions[current_row].height = 30
            current_row += 1

            # 说明行 - 收集所有利润率
            all_profits = []
            for item in items:
                if 'profitPercent' in item:
                    all_profits.append(float(item['profitPercent']))
            unique_profits = set(all_profits)
            if len(unique_profits) == 1:
                profit_str = f"利润率: +{list(unique_profits)[0]}%"
            else:
                profit_str = f"自定义利润率 (多个不同值)"
            ws.merge_cells(f'A{current_row}:C{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = profit_str
            cell.alignment = left_align
            if date and total_cols >= 4:
                date_cell = ws.cell(row=current_row, column=total_cols)
                date_cell.value = date
                date_cell.alignment = Alignment(horizontal='right', vertical='center')
            current_row += 1

            # 表头
            col_display_names = {
                'spec': '规格',
                'ruleName': '类型',
                'originalPrice': '原价',
                'profitPercent': '利润率(%)',
                'calculatedPrice': '价格',
            }

            for col_idx, col_key in enumerate(columns, start=1):
                display_name = col_display_names.get(col_key, col_key)
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = display_name
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

            current_row += 1

            # 写入数据，每行一个item
            for row_idx, item in enumerate(items):
                for col_idx, col_key in enumerate(columns, start=1):
                    value = item.get(col_key, '')

                    # 对价格特殊处理，利润率计算
                    if col_key == 'calculatedPrice' and 'originalPrice' in item and 'profitPercent' in item:
                        original_price = float(item['originalPrice'])
                        profit_percent = float(item['profitPercent'])
                        multiplier = 1 + profit_percent / 100
                        value = round(original_price * multiplier, 2)
                    elif col_key in ['originalPrice', 'calculatedPrice'] and value is not None and value != '':
                        # 保持数值格式
                        try:
                            value = float(value)
                        except:
                            pass
                    elif col_key == 'profitPercent' and value is not None and value != '':
                        try:
                            value = float(value)
                        except:
                            pass

                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = value
                    # 数字右对齐，其他居中
                    if col_key in ['originalPrice', 'profitPercent', 'calculatedPrice']:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = center_align
                    cell.border = thin_border
                    if row_idx % 2 == 1:
                        cell.fill = alt_row_fill

                current_row += 1

            # 自动调整列宽 - 确保内容完全显示
            for col in ws.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                # 使用更大的系数 + 增加最小宽度确保内容完全显示
                adjusted_width = max(max_length * 2 + 2, 8)
                ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # 保存
        wb.save(output_path)

        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=output_filename
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"生成失败: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
