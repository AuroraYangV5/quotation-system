#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
解析Excel报价表，根据设置的利润点生成新报价表
支持单Sheet简单格式和多Sheet复杂格式
输入: 原价表Excel (.xls/.xlsx)
输出: 新报价表Excel
"""

import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import argparse
import sys


class TableData:
    """存储一个表格的数据"""
    def __init__(self, start_row):
        self.spec_col = -1  # 规格所在列
        self.specifications = []  # 规格列表
        self.rules = []  # 规则/商品类型列表
        self.price_data = []  # 价格数据
        self.rule_col_start = 1  # 商品类型起始列
        self.start_row = start_row  # 表格起始行（表头行）


class PriceParser:
    def __init__(self, file_path):
        """
        初始化解析器
        :param file_path: Excel文件路径
        """
        self.file_path = file_path
        self.tables_by_sheet = {}  # {sheet_name: [TableData, ...]}
        self.sheet_titles = {}  # {sheet_name: title}
        self.sheet_dates = {}  # {sheet_name: date}

    def parse_sheet(self, wb, sheet_idx):
        """解析一个sheet"""
        sheet = wb.sheet_by_index(sheet_idx)
        sheet_name = sheet.name
        print(f"\n{'='*60}")
        print(f"正在解析Sheet: [{sheet_name}]，尺寸: {sheet.nrows}行 × {sheet.ncols}列")

        tables = []
        title = None
        date = None

        # 先找标题和日期（通常在前两行）
        for row_idx in range(min(3, sheet.nrows)):
            for col_idx in range(min(5, sheet.ncols)):
                cell_val = sheet.cell_value(row_idx, col_idx)
                if cell_val and '价格' in str(cell_val) and not title:
                    title = str(cell_val).strip()
                if cell_val and str(cell_val).count('.') >= 1 and not date:
                    try:
                        float(str(cell_val))
                        date = str(cell_val)
                    except:
                        pass

        # 如果没找到日期，再扫描一下
        if not date:
            for row_idx in range(min(5, sheet.nrows)):
                for col_idx in range(sheet.ncols):
                    cell_val = sheet.cell_value(row_idx, col_idx)
                    if cell_val and str(cell_val).count('.') >= 1:
                        try:
                            float(str(cell_val))
                            date = str(cell_val)
                            break
                        except:
                            pass
                if date:
                    break

        # 如果没找到标题，用sheet名
        if not title:
            title = sheet_name

        print(f"标题: {title}")
        if date:
            print(f"日期: {date}")

        # 检测表格区域 - 找所有包含表头的行（表头：第一个单元格是"规格"）
        # 规格列通常在第一列，表头行第一个单元格是"规格"
        spec_found = False
        for row_idx in range(0, min(10, sheet.nrows)):
            cell_val = str(sheet.cell_value(row_idx, 0)).strip()
            if '规格' in cell_val or cell_val == '规格':
                # 找到一个表头行
                table = TableData(start_row=row_idx)
                table.spec_col = 0
                spec_found = True

                # 从第二列开始找规则名称，非空即为规则
                for col_idx in range(1, sheet.ncols):
                    cell_val = sheet.cell_value(row_idx, col_idx)
                    cell_val_str = str(cell_val).strip()
                    if cell_val_str and cell_val_str != '':
                        # 检查是否是地址信息，跳过
                        if '地址' in cell_val_str or '电话' in cell_val_str:
                            continue
                        table.rules.append(cell_val_str)

                if len(table.rules) > 0:
                    print(f"  找到表格: 表头行={row_idx+1}，{len(table.rules)}个商品类型")
                    tables.append(table)
                continue

            # 沟槽件特殊情况：副表规格列在第8列（索引7）
            if not spec_found:
                for col_idx in range(1, min(10, sheet.ncols)):
                    cell_val = str(sheet.cell_value(row_idx, col_idx)).strip()
                    if '规格' in cell_val or cell_val == '规格':
                        table = TableData(start_row=row_idx)
                        table.spec_col = col_idx
                        table.rule_col_start = 0  # 规则从开头开始

                        for c_idx in range(0, sheet.ncols):
                            if c_idx == col_idx:
                                continue
                            cell_val = sheet.cell_value(row_idx, c_idx)
                            cell_val_str = str(cell_val).strip()
                            if cell_val_str and cell_val_str != '':
                                if '地址' in cell_val_str or '电话' in cell_val_str:
                                    continue
                                table.rules.append(cell_val_str)

                        if len(table.rules) > 0:
                            print(f"  找到表格: 表头行={row_idx+1}，规格列={col_idx+1}，{len(table.rules)}个商品类型")
                            tables.append(table)
                        break

        # 如果没找到"规格"表头，默认第一列是规格，第一行有内容的是表头
        if not spec_found and len(tables) == 0:
            # 尝试第三行作为表头（复杂表常见）
            # 寻找第一个非空行作为表头
            for row_idx in range(0, min(5, sheet.nrows)):
                has_content = False
                for col_idx in range(1, sheet.ncols):
                    if sheet.cell_value(row_idx, col_idx):
                        has_content = True
                        break
                if has_content:
                    table = TableData(start_row=row_idx)
                    table.spec_col = 0
                    for col_idx in range(1, sheet.ncols):
                        cell_val = sheet.cell_value(row_idx, col_idx)
                        cell_val_str = str(cell_val).strip()
                        if cell_val_str and cell_val_str != '':
                            if '地址' in cell_val_str or '电话' in cell_val_str:
                                continue
                            table.rules.append(cell_val_str)
                    if len(table.rules) > 0:
                        print(f"  自动检测表格: 表头行={row_idx+1}，{len(table.rules)}个商品类型")
                        tables.append(table)
                    break

        # 解析每个表格的数据
        for table in tables:
            # 数据从表头行下一行开始
            start_data_row = table.start_row + 1
            for row_idx in range(start_data_row, sheet.nrows):
                # 跳过全空行
                row_is_empty = True
                for col_idx in range(sheet.ncols):
                    if sheet.cell_value(row_idx, col_idx):
                        row_is_empty = False
                        break
                if row_is_empty:
                    continue

                # 获取规格
                spec_cell = sheet.cell_value(row_idx, table.spec_col)
                spec_str = str(spec_cell).strip()
                # 跳过地址/电话行
                if '地址' in spec_str or '电话' in spec_str or len(spec_str) > 30:
                    continue
                if spec_str == '':
                    continue

                # 收集价格
                prices = []
                price_col_start = table.rule_col_start
                for i, rule in enumerate(table.rules):
                    col_idx = price_col_start + i
                    if col_idx >= sheet.ncols:
                        prices.append(0)
                        continue
                    cell_val = sheet.cell_value(row_idx, col_idx)
                    try:
                        if isinstance(cell_val, float):
                            price = cell_val
                        else:
                            price = float(str(cell_val)) if str(cell_val).strip() else 0
                    except:
                        price = 0
                    prices.append(price)

                # 检查是否全0，全0说明可能这个规格没有数据了
                if all(p == 0 for p in prices):
                    # 可能是空行，跳过
                    continue

                table.specifications.append(spec_str)
                table.price_data.append(prices)

            print(f"  解析完成: {len(table.specifications)}个规格")

        self.tables_by_sheet[sheet_name] = tables
        self.sheet_titles[sheet_name] = title
        self.sheet_dates[sheet_name] = date

        return len(tables) > 0

    def parse(self):
        """
        解析整个Excel文件，支持多sheet
        """
        wb = xlrd.open_workbook(self.file_path)

        print(f"文件包含 {wb.nsheets} 个工作表")

        success = False
        for sheet_idx in range(wb.nsheets):
            if self.parse_sheet(wb, sheet_idx):
                success = True

        # 统计
        total_tables = 0
        total_specs = 0
        for sheet_name, tables in self.tables_by_sheet.items():
            total_tables += len(tables)
            for table in tables:
                total_specs += len(table.specifications)

        print(f"\n{'='*60}")
        print(f"解析完成: 共 {len(self.tables_by_sheet)} 个sheet，{total_tables} 个表格，{total_specs} 个规格")

        return success

    def calculate_new_prices_for_table(self, table, profit_percent):
        """计算单个表格的新价格"""
        multiplier = 1 + profit_percent / 100
        new_prices = []
        for row in table.price_data:
            new_row = [round(p * multiplier, 1) if p > 0 else 0 for p in row]
            new_prices.append(new_row)
        return new_prices

    def generate_new_excel(self, output_path, profit_percent):
        """
        生成新的Excel报价表，保留多sheet结构
        :param output_path: 输出文件路径
        :param profit_percent: 利润率
        """
        # 创建新工作簿
        wb = openpyxl.Workbook()
        # 删除默认创建的sheet
        default_ws = wb.active
        wb.remove(default_ws)

        # 定义样式
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        title_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        alt_row_fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')

        current_row = 1  # 全局行计数器

        for sheet_name, tables in self.tables_by_sheet.items():
            ws = wb.create_sheet(title=sheet_name)
            title = self.sheet_titles.get(sheet_name, sheet_name)
            date = self.sheet_dates.get(sheet_name)

            current_row = 1
            first_table = True

            for table_idx, table in enumerate(tables):
                new_prices = self.calculate_new_prices_for_table(table, profit_percent)

                if first_table:
                    # 第一个表格，写标题
                    total_cols = len(table.rules) + 1  # +1 规格列
                    end_col_letter = openpyxl.utils.get_column_letter(total_cols)
                    ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
                    cell = ws[f'A{current_row}']
                    cell.value = f"{title} 报价表"
                    cell.font = title_font
                    cell.alignment = center_align
                    cell.fill = title_fill
                    ws.row_dimensions[current_row].height = 30
                    current_row += 1

                    # 利润率信息行
                    ws.merge_cells(f'A{current_row}:C{current_row}')
                    cell = ws[f'A{current_row}']
                    cell.value = f"利润率: +{profit_percent}%"
                    cell.alignment = left_align
                    if date:
                        date_cell = ws.cell(row=current_row, column=total_cols)
                        date_cell.value = date
                        date_cell.alignment = Alignment(horizontal='right', vertical='center')
                    current_row += 1

                    # 如果有多个表格，多空一行
                    if len(tables) > 1:
                        current_row += 1

                    first_table = False
                else:
                    # 多个表格之间空一行
                    current_row += 1

                # 表头行
                spec_header_cell = ws.cell(row=current_row, column=1)
                spec_header_cell.value = '规格'
                spec_header_cell.font = header_font
                spec_header_cell.alignment = center_align
                spec_header_cell.border = thin_border
                spec_header_cell.fill = header_fill

                for col_idx, rule in enumerate(table.rules, start=2):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = rule
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    cell.fill = header_fill

                current_row += 1

                # 写入数据
                for row_idx, (spec, prices) in enumerate(zip(table.specifications, new_prices)):
                    # 规格列
                    cell = ws.cell(row=current_row, column=1)
                    cell.value = spec
                    cell.alignment = center_align
                    cell.border = thin_border
                    if row_idx % 2 == 1:
                        cell.fill = alt_row_fill

                    # 价格列
                    for col_idx, price in enumerate(prices, start=2):
                        if price == 0:
                            cell = ws.cell(row=current_row, column=col_idx)
                            cell.value = ''
                            cell.border = thin_border
                            if row_idx % 2 == 1:
                                cell.fill = alt_row_fill
                            continue

                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = price
                        cell.alignment = center_align
                        cell.border = thin_border
                        if row_idx % 2 == 1:
                            cell.fill = alt_row_fill

                    current_row += 1

            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = max_length * 1.5 + 3
                ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = adjusted_width

        # 保存
        wb.save(output_path)
        print(f"\n新报价表已生成: {output_path}")
        print(f"包含Sheet: {list(self.tables_by_sheet.keys())}")


def main():
    parser = argparse.ArgumentParser(description='解析Excel报价表，根据利润率生成新报价表，支持多Sheet复杂格式')
    parser.add_argument('input', help='输入的Excel文件路径，例如: 简单的原价表.xls')
    parser.add_argument('-p', '--profit', type=float, default=20,
                        help='利润率百分比，默认20%，例如: -p 15 表示增加15%')
    parser.add_argument('-o', '--output', help='输出文件路径，默认: 新报价表_利润率{profit}.xlsx')

    args = parser.parse_args()

    # 设置默认输出路径
    if args.output is None:
        if args.profit.is_integer():
            args.output = f"新报价表_利润率{int(args.profit)}%.xlsx"
        else:
            args.output = f"新报价表_利润率{args.profit}%.xlsx"

    try:
        # 解析并生成
        parser = PriceParser(args.input)
        success = parser.parse()
        if not success:
            print("错误: 未能解析到任何表格数据")
            sys.exit(1)
        parser.generate_new_excel(args.output, args.profit)
        print("\n处理完成!")
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
