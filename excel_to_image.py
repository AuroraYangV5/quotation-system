#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel转图片工具
将Excel文件的每个Sheet转换为图片，供GLM大模型视觉识别
支持 .xls 和 .xlsx 格式
"""
import os
from typing import List, Tuple, Callable
from PIL import Image, ImageDraw, ImageFont
import tempfile


def get_excel_reader(excel_path: str):
    """
    根据文件格式获取Excel读取器
    返回: (sheet_names: List[str], get_sheet_data: (idx) -> (nrows, ncols, get_cell))
    """
    ext = os.path.splitext(excel_path)[1].lower()

    if ext == '.xls':
        # .xls 使用xlrd
        import xlrd
        wb = xlrd.open_workbook(excel_path)
        sheet_names = wb.sheet_names()

        def get_sheet_data(sheet_idx):
            sheet = wb.sheet_by_index(sheet_idx)
            nrows = sheet.nrows
            ncols = sheet.ncols
            def get_cell(row, col):
                return sheet.cell_value(row, col)
            return nrows, ncols, get_cell

        return sheet_names, get_sheet_data

    else:
        # .xlsx 使用openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, data_only=True)
        sheet_names = wb.sheetnames

        def get_sheet_data(sheet_idx):
            sheet_name = sheet_names[sheet_idx]
            sheet = wb[sheet_name]
            nrows = sheet.max_row
            ncols = sheet.max_column
            def get_cell(row, col):
                # openpyxl是1-based索引，转成0-based
                return sheet.cell(row=row+1, column=col+1).value
            return nrows, ncols, get_cell

        return sheet_names, get_sheet_data


class ExcelToImageConverter:
    """将Excel Sheet转换为图片"""

    def __init__(self):
        # 配置参数（像素）
        self.cell_width = 120    # 单元格宽度
        self.cell_height = 35   # 单元格高度
        self.header_height = 40 # 表头高度
        self.padding = 10       # 内边距
        self.font_size = 14     # 正文字号
        self.header_font_size = 16 # 表头字号

        # 颜色配置
        self.bg_color = (255, 255, 255)           # 背景色
        self.header_bg_color = (211, 211, 211)    # 表头背景
        self.border_color = (128, 128, 128)       # 边框色
        self.text_color = (0, 0, 0)               # 文字颜色

        # 尝试加载字体
        self.font = None
        self.header_font = None
        self._load_font()

    def _load_font(self):
        """尝试加载中文字体"""
        try:
            if os.name == 'nt':  # Windows
                self.font = ImageFont.truetype("msyh.ttc", self.font_size)
                self.header_font = ImageFont.truetype("msyh.ttc", self.header_font_size)
            else:  # macOS/Linux
                # 尝试常见中文字体
                font_paths = [
                    ("/System/Library/Fonts/PingFang.ttc", "PingFang"),
                    ("/System/Library/Fonts/STHeiti Light.ttc", "STHeiti"),
                    ("/usr/share/fonts/truetype/wqy/wqy-microhei.ttc", "WenQuanYi"),
                    ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", "Noto"),
                    ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "DejaVu"),
                ]
                for path, _ in font_paths:
                    if os.path.exists(path):
                        self.font = ImageFont.truetype(path, self.font_size)
                        self.header_font = ImageFont.truetype(path, self.header_font_size)
                        break
                if self.font is None:
                    self.font = ImageFont.load_default()
                    self.header_font = ImageFont.load_default()
        except Exception:
            self.font = ImageFont.load_default()
            self.header_font = ImageFont.load_default()

    def calculate_image_size(self, max_rows: int, max_cols: int) -> Tuple[int, int]:
        """计算图片需要的尺寸"""
        width = self.padding * 2 + max_cols * self.cell_width
        height = self.padding * 2 + max_rows * self.cell_height
        return width, height

    def draw_cell(self, draw: ImageDraw.Draw, x: int, y: int, w: int, h: int,
                  text: str, font: ImageFont.ImageFont, is_header: bool = False):
        """绘制一个单元格"""
        # 背景
        if is_header:
            draw.rectangle([x, y, x + w, y + h], fill=self.header_bg_color)
        # 边框
        draw.rectangle([x, y, x + w, y + h], outline=self.border_color, width=1)
        # 文字（居中）
        if text:
            text = str(text).strip()
            try:
                # PIL新版本
                bbox = draw.textbbox((0, 0), text, font=font)
                text_w = bbox[2] - bbox[0]
                text_h = bbox[3] - bbox[1]
            except AttributeError:
                # PIL旧版本兼容
                text_w, text_h = draw.textsize(text, font=font)

            text_x = x + (w - text_w) // 2
            text_y = y + (h - text_h) // 2
            draw.text((text_x, text_y), text, fill=self.text_color, font=font)

    def convert_sheet(self, excel_path: str, sheet_idx: int) -> Image.Image:
        """将单个Sheet转换为图片"""
        sheet_names, get_sheet_data = get_excel_reader(excel_path)
        nrows, ncols, get_cell = get_sheet_data(sheet_idx)

        # 计算图片尺寸
        img_width, img_height = self.calculate_image_size(nrows, ncols)
        img = Image.new('RGB', (img_width, img_height), self.bg_color)
        draw = ImageDraw.Draw(img)

        # 绘制每个单元格（0-based索引）
        for row_idx in range(nrows):
            for col_idx in range(ncols):
                x = self.padding + col_idx * self.cell_width
                y = self.padding + row_idx * self.cell_height
                cell_w = self.cell_width
                cell_h = self.cell_height
                cell_val = get_cell(row_idx, col_idx)
                text = cell_val if cell_val is not None else ''
                is_header = row_idx == 0
                font = self.header_font if is_header else self.font
                self.draw_cell(draw, x, y, cell_w, cell_h, text, font, is_header)

        return img

    def convert_all_sheets(self, excel_path: str, output_dir: str = None) -> List[str]:
        """
        将Excel所有Sheet转换为图片
        返回生成的图片文件路径列表
        """
        sheet_names, _ = get_excel_reader(excel_path)

        image_paths = []

        if output_dir is None:
            output_dir = tempfile.gettempdir()

        for idx, sheet_name in enumerate(sheet_names):
            img = self.convert_sheet(excel_path, idx)
            # 保存图片
            safe_name = sheet_name.replace('/', '_').replace('\\', '_')
            img_path = os.path.join(output_dir, f"excel_{idx}_{safe_name}.png")
            img.save(img_path, 'PNG')
            image_paths.append(img_path)

        return image_paths


def convert_excel_to_images(excel_path: str) -> List[str]:
    """便捷函数：将Excel所有Sheet转换为图片"""
    converter = ExcelToImageConverter()
    return converter.convert_all_sheets(excel_path)


__all__ = ['ExcelToImageConverter', 'convert_excel_to_images', 'get_excel_reader']


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        converter = ExcelToImageConverter()
        images = converter.convert_all_sheets(excel_file)
        print(f"生成了 {len(images)} 张图片:")
        for path in images:
            print(f"  {path}")
            print(f"  文件存在: {os.path.exists(path)}")
    else:
        print("用法: python excel_to_image.py <excel文件>")
